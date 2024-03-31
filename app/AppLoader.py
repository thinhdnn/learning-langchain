import os
import glob
from docx import Document
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from  langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import DirectoryLoader, PyPDFLoader
from langchain_community.vectorstores import FAISS
from  langchain_community.embeddings import GPT4AllEmbeddings


class DirectoryLoader:
    def __init__(self, dataPath, extensions):
        self.dataPath = dataPath
        self.extensions = extensions

    def load(self):
        documents = []
        for ext in self.extensions:
            files = glob.glob(os.path.join(self.dataPath, f'*.{ext}'))
            for file in files:
                document = self.load_document(file, ext)
                if document:
                    documents.append(document)
        return documents

    def load_document(self, file_path, ext):
        if ext == 'pdf':
            return self.load_pdf(file_path)
        elif ext == 'docx':
            return self.load_docx(file_path)
        elif ext == 'xlsx':
            return self.load_xlsx(file_path)
        elif ext == 'html':
            return self.load_html(file_path)
        else:
            return None

    def load_pdf(self, file_path):
        # Placeholder for loading PDF files
        pass

    def load_docx(self, file_path):
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text

    def load_xlsx(self, file_path):
        wb = load_workbook(filename=file_path, read_only=True)
        text = ""
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row) + "\n"
        return text

    def load_html(self, file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file, 'html.parser')
            text = soup.get_text()
        return text

def createFromFiles(dataPath, vectorDBPath, extensions):
    # Load data from files
    dirLoader = DirectoryLoader(dataPath, extensions)
    documents = dirLoader.load()

    # Split text into characters
    textSplitters = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
    chunks = textSplitters.split_documents(documents)
    
    # Generate embeddings
    embeddings = GPT4AllEmbeddings(model_file='models/all-MiniLM-L6-v2')
    db = FAISS.from_documents(chunks, embeddings)
    db.save_local(vectorDBPath)

    return db



createFromFiles('data', 'stores/db_faiss', ['pdf', 'docx', 'xlsx', 'html'])